# GUI_App_for_Data_Entry.py
import os
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import messagebox

# Excel file name
filename = "wb.xlsx"

# If Excel file doesn't exist, create it with headers
if not os.path.exists(filename):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "StudentData"
    sheet.append(["Name", "Course", "Semester", "Form No.", "Contact Number", "Email-ID", "Address"])
    wb.save(filename)

# Load the Excel workbook
wb = load_workbook(filename)
sheet = wb.active

# ---------------- Functions ---------------- #

def focus_course(event): course_field.focus_set()
def focus_sem(event): sem_field.focus_set()
def focus_form(event): form_no_field.focus_set()
def focus_contact(event): contact_no_field.focus_set()
def focus_email(event): email_id_field.focus_set()
def focus_address(event): address_field.focus_set()

def clear_fields():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def insert():
    if (not name_field.get() and not course_field.get() and not sem_field.get() and
        not form_no_field.get() and not contact_no_field.get() and
        not email_id_field.get() and not address_field.get()):
        messagebox.showwarning("Empty", "All fields are empty.")
        return

    # Append row and save
    sheet.append([
        name_field.get(),
        course_field.get(),
        sem_field.get(),
        form_no_field.get(),
        contact_no_field.get(),
        email_id_field.get(),
        address_field.get()
    ])
    wb.save(filename)
    messagebox.showinfo("Saved", "Data inserted successfully into wb.xlsx")
    clear_fields()
    name_field.focus_set()

# ---------------- GUI Code ---------------- #
if __name__ == "__main__":
    root = Tk()
    root.title("Student Registration Form")
    root.geometry("520x360")

    heading = Label(root, text="Student Details Form", bg="grey", fg="white", font=("Arial", 14))
    heading.grid(row=0, column=0, columnspan=2, pady=10)

    labels = ["Name", "Course", "Semester", "Form No.", "Contact Number", "Email-ID", "Address"]
    for i, txt in enumerate(labels, start=1):
        Label(root, text=txt, bg="light grey").grid(row=i, column=0, padx=10, pady=5, sticky="w")

    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    name_field.grid(row=1, column=1, ipadx=80, pady=5)
    course_field.grid(row=2, column=1, ipadx=80, pady=5)
    sem_field.grid(row=3, column=1, ipadx=80, pady=5)
    form_no_field.grid(row=4, column=1, ipadx=80, pady=5)
    contact_no_field.grid(row=5, column=1, ipadx=80, pady=5)
    email_id_field.grid(row=6, column=1, ipadx=80, pady=5)
    address_field.grid(row=7, column=1, ipadx=80, pady=5)

    # Bind Enter key to move focus
    name_field.bind("<Return>", focus_course)
    course_field.bind("<Return>", focus_sem)
    sem_field.bind("<Return>", focus_form)
    form_no_field.bind("<Return>", focus_contact)
    contact_no_field.bind("<Return>", focus_email)
    email_id_field.bind("<Return>", focus_address)

    submit = Button(root, text="Submit", fg="Black", bg="blue", command=insert)
    submit.grid(row=8, column=1, pady=15)

    name_field.focus_set()
    root.mainloop()
